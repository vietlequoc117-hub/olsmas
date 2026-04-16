import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import re

def clean_text(text):
    if pd.isna(text):
        return ""
    # Xóa khoảng trắng đầu cuối, chuyển về chữ thường và chuẩn hóa khoảng trắng giữa các từ
    import re
    text = str(text).strip().lower()
    return re.sub(r'\s+', ' ', text)

def get_source_data(file_paths):
    """
    Đọc dữ liệu từ 3 file nguồn và gộp thành một Master List.
    """
    master_data = {}
    
    for path in file_paths:
        if not os.path.exists(path):
            print(f"⚠️ Cảnh báo: Không tìm thấy file {path}")
            continue
            
        print(f"📖 Đang đọc file nguồn: {path}")
        try:
            # Load file để tìm sheet phù hợp
            xl = pd.ExcelFile(path)
            target_sheet = ""
            for sheet in xl.sheet_names:
                if "Bài thi môn Tin học" in sheet:
                    target_sheet = sheet
                    break
            if not target_sheet:
                target_sheet = xl.sheet_names[-1] # Lấy sheet cuối nếu không tìm thấy tên khớp
            
            # Đọc dữ liệu, tìm dòng bắt đầu (STT = 1)
            # Chúng ta đọc toàn bộ và xử lý sau để chính xác hơn
            df_raw = pd.read_excel(path, sheet_name=target_sheet, header=None)
            
            # Tìm dòng có STT = 1 (thường ở cột A - index 0)
            start_row = -1
            for i, row in df_raw.iterrows():
                if str(row[0]).strip() == "1":
                    start_row = i
                    break
            
            if start_row == -1:
                print(f"❌ Lỗi: Không tìm thấy dòng bắt đầu (STT=1) trong sheet {target_sheet}")
                continue
                
            # Đọc lại với header đúng
            df = pd.read_excel(path, sheet_name=target_sheet, skiprows=start_row-1)
            
            # Ánh xạ cột dựa trên vị trí (C=index 2, D=index 3, J=index 9)
            # Lưu ý: pandas read_excel có thể lệch index nếu có cột ẩn, nên ta dùng iloc cho chắc
            for _, row in df.iterrows():
                try:
                    name = clean_text(row.iloc[2]) # Cột C
                    class_name = clean_text(row.iloc[3]) # Cột D
                    score = row.iloc[9] # Cột J
                    
                    if name and class_name:
                        # Key là bộ (Tên, Lớp) để tránh trùng tên khác lớp
                        master_data[(name, class_name)] = score
                except Exception as e:
                    continue
                    
        except Exception as e:
            print(f"❌ Lỗi khi xử lý file {path}: {e}")
            
    return master_data

def sync_to_target(target_file, master_data, output_file):
    """
    Ghi dữ liệu từ Master List vào file đích giữ nguyên định dạng.
    """
    if not os.path.exists(target_file):
        print(f"❌ Lỗi: Không tìm thấy file đích {target_file}")
        return

    print(f"📝 Đang xử lý file đích: {target_file}")
    wb = openpyxl.load_workbook(target_file)
    
    count_updated = 0
    count_not_found = 0
    
    for sheet in wb.worksheets:
        sheet_name = sheet.title
        print(f"  > Đang xử lý sheet: {sheet_name}")
        
        # Trích xuất tên lớp từ tên sheet (Ví dụ: "Tin học 10A1" -> "10A1")
        # Tìm chuỗi có dạng Khối + Lớp (10A1, 11B2...)
        match = re.search(r'(\d{1,2}[A-Z]\d{1,2})', sheet_name)
        if not match:
            # Nếu không tìm thấy bằng regex, thử lấy phần cuối sau dấu cách
            class_from_sheet = clean_text(sheet_name.split()[-1])
        else:
            class_from_sheet = clean_text(match.group(1))
            
        # Duyệt qua các dòng trong sheet đích
        # Giả định dữ liệu bắt đầu từ dòng 5 hoặc tìm dòng có STT=1
        start_row = 1
        for row in range(1, 20):
            if str(sheet.cell(row=row, column=1).value).strip() == "1":
                start_row = row
                break
        
        # Cột C (Họ tên) là cột 3, Cột F (Điểm GK) là cột 6
        # Theo yêu cầu: Họ tên (Cột C), ĐĐG GK (Cột F)
        COL_NAME = 3
        COL_SCORE = 6 # Cột F. Nếu thực tế là cột G thì đổi thành 7
        
        for row in range(start_row, sheet.max_row + 1):
            name_val = sheet.cell(row=row, column=COL_NAME).value
            if not name_val:
                continue
                
            name_clean = clean_text(name_val)
            
            # Thử tìm trong Master List
            score = master_data.get((name_clean, class_from_sheet))
            
            if score is not None:
                sheet.cell(row=row, column=COL_SCORE).value = score
                count_updated += 1
            else:
                # Nếu không thấy, có thể do tên lớp trong file nguồn khác định dạng sheet
                # Log lại để kiểm tra
                if name_clean:
                    # print(f"    ❓ Không tìm thấy điểm cho: {name_val} ({class_from_sheet})")
                    count_not_found += 1

    wb.save(output_file)
    print(f"✅ Hoàn thành! Đã cập nhật {count_updated} học sinh.")
    print(f"ℹ️ Không tìm thấy {count_not_found} học sinh (vui lòng kiểm tra lại tên/lớp).")
    print(f"💾 File kết quả: {output_file}")

if __name__ == "__main__":
    # Cấu hình tên file tại đây
    source_files = ["K10_KetQua.xlsx", "K11_KetQua.xlsx", "K12_KetQua.xlsx"]
    target_file = "BangDiemTongHop.xlsx"
    output_file = "BangDiem_KetQua_DongBo.xlsx"
    
    # Thực hiện
    master_list = get_source_data(source_files)
    if master_list:
        sync_to_target(target_file, master_list, output_file)
    else:
        print("❌ Không có dữ liệu nguồn để đồng bộ.")
